"""Source and LLM adapters used by the AI intake producer."""

from dataclasses import dataclass
from datetime import date, datetime
import io
from typing import Any, Awaitable, Callable, Protocol

import openpyxl
from openpyxl.utils import get_column_letter

from app.intake.models import EvidenceLocator


@dataclass(frozen=True)
class SourceMaterial:
    text: str
    content_type: str
    file_data: bytes | None = None
    deterministic_records: list[dict[str, Any]] | None = None
    evidence_locators: list[EvidenceLocator] | None = None


class SourceCommand(Protocol):
    object_name: str
    content_type: str


class SourceAdapter(Protocol):
    async def read(
        self, command: SourceCommand, max_characters: int
    ) -> SourceMaterial: ...


class LLMAdapter(Protocol):
    async def complete(self, **options: Any) -> str: ...


class IntakeAdapterError(Exception):
    pass


class ProvidedTextAdapter:
    def __init__(self, text: str, content_type: str):
        self._text = text
        self._content_type = content_type

    async def read(self, command: SourceCommand, max_characters: int) -> SourceMaterial:
        return SourceMaterial(
            text=_truncate(self._text, max_characters),
            content_type=self._content_type,
        )


class StoredDocumentAdapter:
    def __init__(
        self,
        download: Callable[[str, str], bytes],
        extract: Callable[[bytes, str], Awaitable[str]],
        bucket: str = "lease-uploads",
    ):
        self._download = download
        self._extract = extract
        self._bucket = bucket

    async def read(self, command: SourceCommand, max_characters: int) -> SourceMaterial:
        try:
            file_data = self._download(self._bucket, command.object_name)
        except Exception as exc:
            raise IntakeAdapterError(f"文件下载失败: {exc}") from exc
        try:
            text = await self._extract(file_data, command.content_type)
        except Exception as exc:
            raise IntakeAdapterError(f"文本提取失败: {exc}") from exc
        return SourceMaterial(
            text=_truncate(text, max_characters),
            content_type=command.content_type,
            file_data=file_data,
        )


class StoredExcelAdapter:
    def __init__(
        self,
        download: Callable[[str, str], bytes],
        bucket: str = "lease-uploads",
    ):
        self._download = download
        self._bucket = bucket

    async def read(self, command: SourceCommand, max_characters: int) -> SourceMaterial:
        try:
            file_data = self._download(self._bucket, command.object_name)
        except Exception as exc:
            raise IntakeAdapterError(f"文件下载失败: {exc}") from exc
        try:
            text, records, locators = _read_excel_contracts(file_data)
        except Exception as exc:
            raise IntakeAdapterError(f"Excel 文本展开失败: {exc}") from exc
        return SourceMaterial(
            text=_truncate(text, max_characters),
            content_type=command.content_type,
            file_data=file_data,
            deterministic_records=records or None,
            evidence_locators=locators or None,
        )


class RuntimeLLMAdapter:
    def __init__(self, client: Any):
        self._client = client

    async def complete(self, **options: Any) -> str:
        system = str(options.pop("system"))
        prompt = str(options.pop("prompt"))
        response = await self._client.chat_completion(
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": prompt},
            ],
            **options,
        )
        return response["choices"][0]["message"]["content"]


def _truncate(text: str, max_characters: int) -> str:
    if len(text) <= max_characters:
        return text
    return text[:max_characters] + "\n... (truncated)"


def _read_excel_contracts(
    file_data: bytes, max_rows: int = 200, max_cols: int = 60
) -> tuple[str, list[dict[str, Any]], list[EvidenceLocator]]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
    try:
        text_parts = [
            "Excel workbook cell dump for AI semantic parsing.",
            "The cell coordinates are source evidence. Infer headers and fields from nearby cells, sheet names, and row context.",
        ]
        records: list[dict[str, Any]] = []
        locators: list[EvidenceLocator] = []
        for sheet in workbook.worksheets:
            text_parts.append(f"\n## Sheet: {sheet.title}")
            for row in sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row or 0, max_rows),
                min_col=1,
                max_col=min(sheet.max_column or 0, max_cols),
            ):
                cells = [
                    f"{cell.coordinate}={_format_excel_value(cell.value)}"
                    for cell in row
                    if cell.value is not None and str(cell.value).strip()
                ]
                if cells:
                    text_parts.append(f"Row {row[0].row}: " + " | ".join(cells))

            values = list(sheet.iter_rows(values_only=True))
            header_index = None
            indexes: dict[str, int] = {}
            for index, row in enumerate(values[:20]):
                candidate = _header_indexes(list(row))
                if "contract_number" in candidate and "lessor" in candidate:
                    header_index, indexes = index, candidate
                    break
            if header_index is None:
                continue
            for row_number, row in enumerate(
                values[header_index + 1 :], start=header_index + 2
            ):
                contract_number = str(
                    _cell(row, indexes, "contract_number") or ""
                ).strip()
                if not contract_number or contract_number.lower() in {"none", "null"}:
                    continue
                lease_scope = (
                    str(_cell(row, indexes, "lease_scope") or "in_scope").strip()
                    or "in_scope"
                )
                records.append(
                    {
                        "contract_number": contract_number,
                        "contract_name": str(
                            _cell(row, indexes, "contract_name") or contract_number
                        ).strip(),
                        "lessee": str(_cell(row, indexes, "lessee") or "").strip(),
                        "lessor": str(_cell(row, indexes, "lessor") or "").strip(),
                        "store_name": str(
                            _cell(row, indexes, "store_name") or ""
                        ).strip(),
                        "store_address": str(
                            _cell(row, indexes, "store_address") or ""
                        ).strip(),
                        "commencement_date": _format_excel_value(
                            _cell(row, indexes, "commencement_date")
                        ),
                        "lease_start_date": _format_excel_value(
                            _cell(row, indexes, "lease_start_date")
                            or _cell(row, indexes, "commencement_date")
                        ),
                        "lease_end_date": _format_excel_value(
                            _cell(row, indexes, "lease_end_date")
                        ),
                        "currency": str(_cell(row, indexes, "currency") or "").strip(),
                        "asset_type": _cell(row, indexes, "asset_type"),
                        "fixed_rent_amount": _cell(row, indexes, "fixed_rent_amount"),
                        "payment_frequency": "monthly",
                        "payment_timing": _cell(row, indexes, "payment_timing"),
                        "renewal_option": _cell(row, indexes, "renewal_option"),
                        "termination_option": _cell(row, indexes, "termination_option"),
                        "discount_rate_type": str(
                            _cell(row, indexes, "discount_rate_type") or ""
                        ).strip(),
                        "discount_rate": _cell(row, indexes, "discount_rate"),
                        "lease_scope": lease_scope,
                        "suggested_scope": lease_scope,
                        "scope_source": "ledger",
                        "scope_confidence": 0.9,
                        "confidence": 0.9,
                    }
                )
                record_index = len(records) - 1
                locators.append(
                    EvidenceLocator(
                        field=f"contracts[{record_index}]",
                        source=(
                            f"{sheet.title}!A{row_number}:"
                            f"{get_column_letter(len(row))}{row_number}"
                        ),
                        quote=contract_number,
                    )
                )
        return "\n".join(text_parts), records, locators
    finally:
        workbook.close()


def _format_excel_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text


def _header_indexes(headers: list[Any]) -> dict[str, int]:
    aliases = {
        "contract_number": ["合同编号", "contract_number", "合同号"],
        "contract_name": ["合同名称", "contract_name", "合同名"],
        "lessee": ["承租方", "法人主体", "legal_entity", "lessee"],
        "lessor": ["出租方", "lessor"],
        "store_name": ["门店/资产名称", "门店名称", "资产名称", "store_name"],
        "store_address": ["门店/资产地址", "门店地址", "资产地址", "store_address"],
        "asset_type": ["资产类型", "资产类别", "asset_type", "asset_category"],
        "currency": ["币种", "currency"],
        "commencement_date": [
            "起租日(commencement)",
            "起租日",
            "commencement_date",
            "租赁起始日",
        ],
        "lease_start_date": ["租赁开始日", "lease_start_date"],
        "lease_end_date": ["租赁结束日", "租期结束日", "lease_end_date"],
        "renewal_option": ["续租选择权", "renewal_option"],
        "termination_option": ["终止选择权判断", "终止选择权", "termination_option"],
        "fixed_rent_amount": ["月租金", "固定租金", "fixed_rent_amount"],
        "payment_timing": ["付款时点", "payment_timing"],
        "discount_rate": ["折现率", "discount_rate"],
        "discount_rate_type": ["折现率类型", "discount_rate_type"],
        "lease_scope": ["范围判定(lease_scope)", "lease_scope", "范围判定"],
    }
    normalized = {
        str(header or "").strip().lower(): index for index, header in enumerate(headers)
    }
    result: dict[str, int] = {}
    for field, names in aliases.items():
        for name in names:
            key = name.strip().lower()
            if key in normalized:
                result[field] = normalized[key]
                break
    return result


def _cell(row: tuple[Any, ...], indexes: dict[str, int], field: str) -> Any:
    index = indexes.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]
