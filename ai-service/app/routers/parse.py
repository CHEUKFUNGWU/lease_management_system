from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import date, datetime
import io
import json
import logging

from app.services.llm import llm_client
from app.services.storage import download_from_minio, get_minio_client
from app.services.document_extractor import extract_text

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

router = APIRouter()
logger = logging.getLogger(__name__)


class ParseRequest(BaseModel):
    file_id: str
    file_type: str  # contract, payment_schedule, event
    task_type: Optional[str] = "contract_extraction"


class ParseResponse(BaseModel):
    task_id: str
    file_id: str
    status: str
    extracted_data: Optional[dict] = None
    confidence_scores: Optional[dict] = None
    warnings: Optional[List[str]] = None


class ContractDraftRequest(BaseModel):
    file_id: str
    object_name: str  # MinIO object name for file download
    content_type: str = "application/pdf"  # MIME type for text extraction
    file_content: Optional[str] = None  # Optional: pre-extracted text (if provided, skip extraction)
    mode: str = "assist"  # "assist" or "auto-post"


class ContractDraftResponse(BaseModel):
    task_id: str
    file_id: str
    mode: str
    draft_type: str = "contract_draft"
    status: str
    extracted_data: Dict[str, Any]
    confidence_scores: Dict[str, float]
    missing_fields: List[str]
    warnings: List[str]
    requires_human_confirmation: bool


def _check_discount_rate_missing(extracted: dict) -> tuple[bool, list]:
    """检查折现率是否缺失"""
    missing = []
    warnings = []
    
    if not extracted.get("discount_rate_type") and not extracted.get("discount_rate"):
        missing.append("discount_rate")
        warnings.append("【关键】合同缺少折现率信息。AI 不得猜测折现率，需要人工确认。")
        warnings.append("建议处理方式：")
        warnings.append("1. 检查合同文本中是否明确提到利率")
        warnings.append("2. 从系统政策库中查找适用的 IBR")
        warnings.append("3. 按法人/租期区间/门店类型匹配利率政策")
        warnings.append("4. 如无法唯一确定，请人工输入或选择")
    
    return len(missing) > 0, warnings


def _check_currency_missing(extracted: dict) -> tuple[bool, list]:
    """检查货币是否缺失 — AI 不得猜测货币，必须询问用户确认"""
    missing = False
    warnings = []
    currency = extracted.get("currency")
    
    # Treat empty, null, or placeholder values as missing
    if not currency or str(currency).strip() == "" or str(currency).lower() in ("unknown", "null", "none"):
        missing = True
        warnings.append("【必须确认】AI 未识别到合同货币。根据 IFRS 16 计量要求，货币直接影响租赁负债现值计算和后续会计分录。")
        warnings.append("请在上传后手动选择货币（CNY / USD / EUR 等）。AI 不会猜测货币。")
    
    return missing, warnings


def _check_critical_fields(extracted: dict) -> tuple[list, list]:
    """检查关键字段是否缺失或低置信度"""
    missing = []
    warnings = []
    
    critical_fields = [
        ("contract_number", "合同编号"),
        ("lessee", "承租方"),
        ("lessor", "出租方"),
        ("commencement_date", "租赁起始日"),
        ("lease_start_date", "租赁开始日"),
        ("lease_end_date", "租期结束日"),
        ("fixed_rent_amount", "固定租金金额"),
        ("payment_timing", "付款时点（先付/后付）"),
    ]
    
    for field, label in critical_fields:
        if not extracted.get(field):
            missing.append(field)
            warnings.append(f"【关键字段缺失】{label}({field}) 未识别到，必须人工补充")
    
    return missing, warnings


def _normalize_lease_scope(extracted: dict) -> tuple[str, list]:
    """Normalize AI suggested IFRS 16 scope and emit review warnings."""
    warnings = []
    scope = str(extracted.get("suggested_scope") or extracted.get("lease_scope") or "in_scope").strip()
    allowed = {"in_scope", "short_term_exempt", "low_value_exempt", "not_a_lease"}
    if scope not in allowed:
        warnings.append("【范围判定】AI 未能给出有效 lease_scope，默认按 in_scope 进入人工确认。")
        scope = "in_scope"

    confidence = extracted.get("scope_confidence")
    try:
        confidence_value = float(confidence) if confidence is not None and confidence != "" else None
    except (TypeError, ValueError):
        confidence_value = None

    if confidence_value is None or confidence_value < 0.8:
        warnings.append("【必须确认】租赁范围判定置信度不足，需要人工确认是否资本化、短期/低价值豁免或非租赁。")

    extracted["lease_scope"] = scope
    extracted["suggested_scope"] = scope
    extracted["scope_source"] = "ai_suggested"
    if confidence_value is not None:
        extracted["scope_confidence"] = confidence_value
    return scope, warnings


def _sanitize_confidence_scores(scores: dict) -> dict[str, float]:
    """将 LLM 返回的置信度统一清洗为 float，None/空值转为 0.0。"""
    sanitized: dict[str, float] = {}
    for key, value in (scores or {}).items():
        if value is None or value == "":
            sanitized[key] = 0.0
            continue
        try:
            sanitized[key] = float(value)
        except (TypeError, ValueError):
            sanitized[key] = 0.0
    return sanitized


@router.post("/parse", response_model=ParseResponse)
async def parse_document(request: ParseRequest):
    """
    解析文档并抽取结构化字段
    """
    return {
        "task_id": "task_" + request.file_id,
        "file_id": request.file_id,
        "status": "pending",
        "extracted_data": None,
        "confidence_scores": None,
        "warnings": ["AI Assist Mode: 识别结果需人工确认后入库"]
    }


@router.post("/parse/contract", response_model=ContractDraftResponse)
async def parse_contract(request: ContractDraftRequest):
    """
    解析合同文件，生成合同草稿
    
    默认 Assist Mode：
    - AI 只生成草稿，不直接写入正式台账
    - 所有结果需 Editor 确认
    - 关键字段缺失时必须标记人工确认
    """
    
    if request.mode != "assist":
        raise HTTPException(
            status_code=400, 
            detail="当前仅支持 Assist Mode。Auto-Post Mode 需另行配置"
        )
    
    # Extract text: use provided file_content or download from MinIO + PaddleOCR
    file_content = request.file_content
    if not file_content:
        # Download file from MinIO
        try:
            file_data = download_from_minio("lease-uploads", request.object_name)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"文件下载失败: {str(e)}")
        
        # Extract text using PaddleOCR (primary) + fallback
        try:
            file_content = await extract_text(file_data, request.content_type)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"文本提取失败: {str(e)}")
        
        if len(file_content) > 15000:
            file_content = file_content[:15000] + "\n... (truncated)"
    
    prompt = f"""
    你是一位专业的 IFRS 16 租赁合同解析专家。请从以下合同文本中提取关键字段。
    
    【重要规则 — 必须遵守】
    1. 如果合同中未明确提到折现率/利率，不要猜测，标记为缺失
    2. 如果合同中未明确提到货币（币种），不要猜测，标记为缺失或 unknown
    3. 区分先付租金(prepaid)和后付租金(postpaid)
    4. 区分固定租金和变量租金
    5. 识别租赁成分和非租赁成分(CAM、服务费)
    6. 承租方(lessee)是租赁合同的乙方，即使用物业并支付租金的一方
    7. 出租方(lessor)是租赁合同的甲方，即提供物业并收取租金的一方
    8. 必须做 IFRS 16 范围初判：是否存在已识别资产、承租方是否控制使用、租期是否 ≤12 个月、是否低价值资产。AI 只能建议，不能直接入正式账。
    
    合同文本:
    {file_content}
    
    请提取以下字段（JSON 格式）:
    - contract_number: 合同编号
    - contract_name: 合同名称
    - lessee: 承租方名称（合同中承租人/乙方对应的完整公司名称）
    - lessor: 出租方名称（合同中出租人/甲方对应的完整公司名称）
    - store_name: 门店/物业名称（如有明确提及，否则留空）
    - store_address: 门店/物业地址（如有明确提及，否则留空）
    - commencement_date: 租赁起始日 (YYYY-MM-DD)
    - lease_start_date: 租赁开始日 (YYYY-MM-DD)
    - lease_end_date: 租期结束日 (YYYY-MM-DD)
    - currency: 币种 (CNY/USD/EUR)。如果合同中没有明确提到货币，返回 null 或空字符串，不要猜测
    - asset_type: 标的资产类型 "real_estate" | "vehicle" | "it_equipment" | "machinery" | "other"
    - fixed_rent_amount: 固定租金金额（仅数字，不含货币单位）
    - payment_frequency: 付款频率 (monthly/quarterly/yearly)
    - payment_timing: 付款时点 (prepaid/postpaid)。如果合同写明"每月X日前支付"则为prepaid；"每月X日后支付"或"月末支付"则为postpaid
    - renewal_option: 是否有续租选择权 (true/false)
    - termination_option: 是否有终止选择权 (true/false)
    - cam_amount: 物业管理费 (如有，仅数字)
    - service_fee: 服务费 (如有，仅数字)
    - discount_rate_type: 折现率类型 (如合同中提及)
    - discount_rate: 折现率数值 (如合同中提及)
    - is_lease: 是否构成 IFRS 16 租赁 (true/false)
    - suggested_scope: "in_scope" | "short_term_exempt" | "low_value_exempt" | "not_a_lease"
    - exemption_reason: 范围判定依据，如"租期 10 个月且无续租意图"、"未识别特定资产"
    - scope_confidence: 范围判定置信度 (0-1)
    
    请以 JSON 格式输出，包含:
    - extracted_fields: 提取的字段
    - confidence_scores: 每个字段的置信度 (0-1)
    - overall_confidence: 总体置信度
    - missing_fields: 识别为缺失的字段列表
    - warnings: 需要人工注意的问题列表
    """
    
    try:
        response = await llm_client.chat_completion(
            messages=[
                {"role": "system", "content": "你是一位专业的 IFRS 16 租赁合同解析专家。请准确提取合同字段。如果字段未在合同中出现，不要猜测，标记为缺失。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=2500,
            response_format={"type": "json_object"}
        )
        
        content = response["choices"][0]["message"]["content"]
        
        # Parse JSON response (simplified for MVP)
        import json
        try:
            parsed = json.loads(content)
        except:
            parsed = {
                "extracted_fields": {},
                "confidence_scores": {},
                "overall_confidence": 0.5,
                "missing_fields": [],
                "warnings": ["解析响应格式异常，请人工检查"]
            }
        
        extracted = parsed.get("extracted_fields", {})
        confidence = _sanitize_confidence_scores(parsed.get("confidence_scores", {}))
        
        # Check for discount rate missing
        dr_missing, dr_warnings = _check_discount_rate_missing(extracted)
        
        # Check for currency missing — AI must not guess currency
        currency_missing, currency_warnings = _check_currency_missing(extracted)
        
        # Check for critical fields
        missing_fields, field_warnings = _check_critical_fields(extracted)

        # IFRS 16 scope gate suggestion
        _, scope_warnings = _normalize_lease_scope(extracted)
        
        all_warnings = dr_warnings + currency_warnings + field_warnings + scope_warnings + parsed.get("warnings", [])
        
        # Determine if human confirmation is required
        requires_human = (
            dr_missing or
            currency_missing or
            len(missing_fields) > 0 or
            (extracted.get("scope_confidence") is None or float(extracted.get("scope_confidence", 0)) < 0.8) or
            parsed.get("overall_confidence", 1.0) < 0.8
        )
        
        return {
            "task_id": "task_" + request.file_id,
            "file_id": request.file_id,
            "mode": "assist",
            "draft_type": "contract_draft",
            "status": "draft_generated",
            "extracted_data": extracted,
            "confidence_scores": confidence,
            "missing_fields": missing_fields + (["discount_rate"] if dr_missing else []) + (["currency"] if currency_missing else []),
            "warnings": all_warnings,
            "requires_human_confirmation": requires_human
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")


class PaymentScheduleParseRequest(BaseModel):
    file_id: str
    object_name: str
    content_type: str
    mode: str = "assist"


class PaymentScheduleItem(BaseModel):
    period_start: str
    period_end: str
    due_date: str
    amount: float
    payment_timing: str
    is_fixed: bool
    is_lease_component: bool
    amount_type: Optional[str] = "fixed_rent"
    currency: Optional[str] = "CNY"
    confidence: Optional[float] = 0.9


class PaymentScheduleParseResponse(BaseModel):
    task_id: str
    file_id: str
    mode: str
    draft_type: str
    status: str
    schedules: List[PaymentScheduleItem]
    confidence_scores: Dict[str, float]
    missing_fields: List[str]
    warnings: List[str]
    requires_human_confirmation: bool


def _normalize_header(value: str) -> str:
    value = value.strip().lower().replace(" ", "_").replace("-", "_")
    aliases = {
        "period_start": "period_start",
        "覆盖期间起始日": "period_start",
        "期间开始": "period_start",
        "period_end": "period_end",
        "覆盖期间结束日": "period_end",
        "期间结束": "period_end",
        "due_date": "due_date",
        "应付日期": "due_date",
        "应付日": "due_date",
        "amount": "amount",
        "金额": "amount",
        "payment_timing": "payment_timing",
        "付款时点": "payment_timing",
        "amount_type": "amount_type",
        "金额类型": "amount_type",
        "currency": "currency",
        "币种": "currency",
        "component": "component",
        "成分": "component",
        "is_fixed": "is_fixed",
        "固定租金": "is_fixed",
        "is_lease_component": "is_lease_component",
        "租赁成分": "is_lease_component",
    }
    return aliases.get(value, value)


def _truthy_cell(value: str, default: bool) -> bool:
    normalized = value.strip().lower()
    if normalized in ("true", "yes", "y", "1", "是", "租赁", "lease"):
        return True
    if normalized in ("false", "no", "n", "0", "否", "非租赁", "non_lease", "non-lease"):
        return False
    return default


def _fallback_parse_payment_schedule_text(file_content: str, reason: str) -> Dict[str, Any]:
    """Best-effort Office table fallback used only when LLM parsing is unavailable."""
    rows: List[List[str]] = []
    for raw_line in file_content.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("###") or set(line.replace("|", "").replace("-", "").strip()) == set():
            continue
        if "|" not in line:
            continue
        cells = [cell.strip() for cell in line.split("|")]
        if len(cells) >= 3:
            rows.append(cells)

    if not rows:
        return {
            "schedules": [],
            "overall_confidence": 0.0,
            "missing_fields": ["all"],
            "warnings": [f"LLM 解析不可用，且 Office 表格兜底未找到可读行: {reason}"],
            "total_schedules": 0,
        }

    header = [_normalize_header(cell) for cell in rows[0]]
    required = {"due_date", "amount"}
    if not required.issubset(set(header)):
        return {
            "schedules": [],
            "overall_confidence": 0.0,
            "missing_fields": sorted(required - set(header)),
            "warnings": [f"LLM 解析不可用，Office 表格兜底无法识别必要列: {reason}"],
            "total_schedules": 0,
        }

    schedules = []
    for raw_cells in rows[2:] if len(rows) > 1 and all(cell.replace("-", "").strip() == "" for cell in rows[1]) else rows[1:]:
        row = {header[i]: raw_cells[i] if i < len(raw_cells) else "" for i in range(len(header))}
        if not row.get("due_date") or not row.get("amount"):
            continue
        try:
            amount = float(str(row["amount"]).replace(",", ""))
        except ValueError:
            continue
        amount_type = row.get("amount_type") or "fixed_rent"
        component = row.get("component", "")
        is_fixed = _truthy_cell(row.get("is_fixed", ""), amount_type not in ("turnover_rent", "variable_rent"))
        is_lease_component = _truthy_cell(row.get("is_lease_component", ""), component != "non_lease" and amount_type not in ("cam", "service_fee"))
        schedules.append({
            "period_start": row.get("period_start") or row["due_date"],
            "period_end": row.get("period_end") or row["due_date"],
            "due_date": row["due_date"],
            "amount": amount,
            "payment_timing": row.get("payment_timing") or "postpaid",
            "is_fixed": is_fixed,
            "is_lease_component": is_lease_component,
            "amount_type": amount_type,
            "currency": row.get("currency") or "",
            "confidence": 0.65,
        })

    return {
        "schedules": schedules,
        "overall_confidence": 0.65 if schedules else 0.0,
        "missing_fields": [] if schedules else ["all"],
        "warnings": [f"LLM 解析不可用，已使用 Office 表格兜底读取，必须人工复核: {reason}"],
        "total_schedules": len(schedules),
    }


@router.post("/parse/payment-schedule", response_model=PaymentScheduleParseResponse)
async def parse_payment_schedule(request: PaymentScheduleParseRequest):
    """
    解析租金表文件，提取付款计划草稿
    
    流程:
    1. 从 MinIO 下载文件
    2. 提取文本 (PDF/Excel)
    3. LLM 解析付款计划
    4. 返回结构化草稿 + 置信度
    """
    
    if request.mode != "assist":
        raise HTTPException(
            status_code=400,
            detail="当前仅支持 Assist Mode。Auto-Post Mode 需另行配置"
        )
    
    # Download file from MinIO
    try:
        file_data = download_from_minio("lease-uploads", request.object_name)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件下载失败: {str(e)}")
    
    # Extract text
    try:
        file_content = await extract_text(file_data, request.content_type)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文本提取失败: {str(e)}")
    
    if len(file_content) > 15000:
        file_content = file_content[:15000] + "\n... (truncated)"
    
    prompt = f"""
你是一位专业的 IFRS 16 租金表解析专家。请从以下租金表内容中提取付款计划信息。

【重要规则 - 必须遵守】
1. 每笔付款必须识别：先付(prepaid)还是后付(postpaid)
   - 先付：在覆盖期间开始前支付（如月初预付当月租金）
   - 后付：在覆盖期间结束后支付（如月末支付当月租金）
2. 区分固定租金和变量租金（turnover rent / sales-based rent 必须标记为变量）
3. 区分租赁成分和非租赁成分（CAM、管理费、服务费等）
4. 金额必须是数字，不要包含货币符号
5. 日期格式必须为 YYYY-MM-DD
6. 如果租金表是月度数据，期间起始日=当月1日，期间结束日=当月最后一日
7. 如果某期金额为空或为0，跳过该行

租金表内容:
{file_content}

请提取每笔付款，以 JSON 数组格式输出。每个元素包含:
- period_start: 覆盖期间起始日 (YYYY-MM-DD)
- period_end: 覆盖期间结束日 (YYYY-MM-DD)
- due_date: 应付日期 (YYYY-MM-DD)
- amount: 金额 (纯数字)
- payment_timing: "prepaid" 或 "postpaid"
- is_fixed: true/false
- is_lease_component: true/false
- amount_type: "fixed_rent" | "turnover_rent" | "cam" | "service_fee" | "tax" | "deposit" | "other"
- currency: "CNY" 或文件中出现的币种。如果文件中未明确提到货币，返回 null 或空字符串，不要猜测
- confidence: 该笔识别的置信度 (0.0-1.0)

额外输出字段（JSON 对象顶层）:
- overall_confidence: 总体置信度 (0.0-1.0)
- missing_fields: 识别中遇到问题的字段列表
- warnings: 需要人工注意的问题列表（如：日期格式不确定、金额可能有误等）
- total_schedules: 识别到的付款笔数

请以纯 JSON 格式输出，不要包含任何 markdown 代码块标记。
"""
    
    try:
        response = await llm_client.chat_completion(
            messages=[
                {"role": "system", "content": "你是一位专业的 IFRS 16 租金表解析专家。请准确提取付款计划信息，严格遵守日期格式和金额格式要求。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=4000
        )
        
        content = response["choices"][0]["message"]["content"]
        
        # Clean up markdown code blocks if present
        content = content.strip()
        if content.startswith("```json"):
            content = content[7:]
        if content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
        content = content.strip()
        
        # Parse JSON
        try:
            parsed = json.loads(content)
            # Handle case where LLM returns a JSON array directly
            if isinstance(parsed, list):
                parsed = {
                    "schedules": parsed,
                    "overall_confidence": 0.8,
                    "missing_fields": [],
                    "warnings": [],
                    "total_schedules": len(parsed)
                }
        except json.JSONDecodeError:
            # Try to extract JSON array from text
            try:
                start = content.find("[")
                end = content.rfind("]") + 1
                if start >= 0 and end > start:
                    arr = json.loads(content[start:end])
                    parsed = {"schedules": arr, "overall_confidence": 0.7, "missing_fields": [], "warnings": ["JSON 解析使用了启发式提取"], "total_schedules": len(arr)}
                else:
                    parsed = {"schedules": [], "overall_confidence": 0.0, "missing_fields": ["all"], "warnings": ["无法解析 LLM 输出"], "total_schedules": 0}
            except Exception:
                parsed = {"schedules": [], "overall_confidence": 0.0, "missing_fields": ["all"], "warnings": ["无法解析 LLM 输出"], "total_schedules": 0}
        
        schedules = parsed.get("schedules", [])
        
        # Validate and normalize schedules
        validated_schedules = []
        warnings = parsed.get("warnings", [])
        
        for i, s in enumerate(schedules):
            # Validate required fields
            if not s.get("due_date") or not s.get("amount"):
                warnings.append(f"第 {i+1} 行缺少必要字段 (due_date/amount)，已跳过")
                continue
            
            # Validate date format
            for date_field in ["period_start", "period_end", "due_date"]:
                if s.get(date_field) and len(s[date_field]) != 10:
                    warnings.append(f"第 {i+1} 行 {date_field} 日期格式可能不正确: {s[date_field]}")
            
            # Validate amount is numeric
            try:
                amount = float(s["amount"])
                if amount <= 0:
                    warnings.append(f"第 {i+1} 行金额 <= 0，已跳过")
                    continue
            except (ValueError, TypeError):
                warnings.append(f"第 {i+1} 行金额无法解析为数字: {s.get('amount')}")
                continue
            
            validated_schedules.append(PaymentScheduleItem(
                period_start=s.get("period_start", s.get("due_date", "")),
                period_end=s.get("period_end", s.get("due_date", "")),
                due_date=s["due_date"],
                amount=amount,
                payment_timing=s.get("payment_timing", "postpaid"),
                is_fixed=s.get("is_fixed", True),
                is_lease_component=s.get("is_lease_component", True),
                amount_type=s.get("amount_type", "fixed_rent"),
                currency=s.get("currency", "CNY"),
                confidence=s.get("confidence", 0.8),
            ))
        
        # Check for low confidence items
        low_confidence_count = sum(1 for s in validated_schedules if (s.confidence or 1.0) < 0.8)
        if low_confidence_count > 0:
            warnings.append(f"有 {low_confidence_count} 笔付款的置信度低于 0.8，建议人工复核")
        
        # Check for prepaid vs postpaid consistency
        timings = set(s.payment_timing for s in validated_schedules)
        if len(timings) > 1:
            warnings.append("租金表中同时出现先付和后付，请确认是否正确")
        
        overall_confidence = parsed.get("overall_confidence", 0.8)
        requires_human = (
            overall_confidence < 0.8 or
            len(validated_schedules) == 0 or
            low_confidence_count > len(validated_schedules) / 2
        )
        
        return {
            "task_id": "task_ps_" + request.file_id,
            "file_id": request.file_id,
            "mode": "assist",
            "draft_type": "payment_schedule_draft",
            "status": "draft_generated",
            "schedules": validated_schedules,
            "confidence_scores": {
                "overall": overall_confidence,
                "average_item": sum((s.confidence or 0.8) for s in validated_schedules) / max(len(validated_schedules), 1),
            },
            "missing_fields": parsed.get("missing_fields", []),
            "warnings": warnings + ["AI Assist Mode: 付款计划草稿需人工确认后入库"],
            "requires_human_confirmation": requires_human,
        }
    except Exception as e:
        reason = f"{type(e).__name__}: {str(e) or repr(e)}"
        logger.exception("LLM payment schedule parsing failed; using Office table fallback")
        parsed = _fallback_parse_payment_schedule_text(file_content, reason)
        validated_schedules = []
        warnings = parsed.get("warnings", [])
        for i, s in enumerate(parsed.get("schedules", [])):
            try:
                amount = float(s["amount"])
            except (ValueError, TypeError, KeyError):
                warnings.append(f"兜底读取第 {i+1} 行金额无法解析，已跳过")
                continue
            validated_schedules.append(PaymentScheduleItem(
                period_start=s.get("period_start", s.get("due_date", "")),
                period_end=s.get("period_end", s.get("due_date", "")),
                due_date=s.get("due_date", ""),
                amount=amount,
                payment_timing=s.get("payment_timing", "postpaid"),
                is_fixed=s.get("is_fixed", True),
                is_lease_component=s.get("is_lease_component", True),
                amount_type=s.get("amount_type", "fixed_rent"),
                currency=s.get("currency", ""),
                confidence=s.get("confidence", 0.65),
            ))

        return {
            "task_id": "task_ps_" + request.file_id,
            "file_id": request.file_id,
            "mode": "assist",
            "draft_type": "payment_schedule_draft",
            "status": "draft_generated",
            "schedules": validated_schedules,
            "confidence_scores": {
                "overall": parsed.get("overall_confidence", 0.0),
                "average_item": sum((s.confidence or 0.65) for s in validated_schedules) / max(len(validated_schedules), 1),
            },
            "missing_fields": parsed.get("missing_fields", []),
            "warnings": warnings + ["AI Assist Mode: 付款计划草稿需人工确认后入库"],
            "requires_human_confirmation": True,
        }


class ContractBatchDraftRequest(BaseModel):
    file_id: str
    object_name: str
    content_type: str = "application/pdf"
    file_content: Optional[str] = None
    mode: str = "assist"


class ContractBatchItem(BaseModel):
    contract_number: str
    contract_name: str
    lessee: str
    lessor: str
    store_name: Optional[str] = ""
    store_address: Optional[str] = ""
    commencement_date: str
    lease_start_date: str
    lease_end_date: str
    currency: Optional[str] = ""
    asset_type: Optional[str] = "real_estate"
    fixed_rent_amount: Optional[float] = 0.0
    payment_frequency: Optional[str] = "monthly"
    payment_timing: Optional[str] = "postpaid"
    renewal_option: Optional[bool] = False
    termination_option: Optional[bool] = False
    cam_amount: Optional[float] = 0.0
    service_fee: Optional[float] = 0.0
    discount_rate_type: Optional[str] = ""
    discount_rate: Optional[float] = 0.0
    is_lease: Optional[bool] = True
    lease_scope: Optional[str] = "in_scope"
    suggested_scope: Optional[str] = "in_scope"
    exemption_reason: Optional[str] = ""
    scope_source: Optional[str] = "ai_suggested"
    scope_confidence: Optional[float] = 0.0
    confidence: Optional[float] = 0.8
    missing_fields: Optional[List[str]] = []
    warnings: Optional[List[str]] = []


class ContractBatchDraftResponse(BaseModel):
    task_id: str
    file_id: str
    mode: str
    draft_type: str = "contract_batch_draft"
    status: str
    contracts: List[ContractBatchItem]
    total_count: int
    confidence_scores: Dict[str, float]
    missing_fields: List[str]
    warnings: List[str]
    requires_human_confirmation: bool


def _format_excel_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _normalize_payment_timing(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in ("prepaid", "advance", "先付", "期初", "预付"):
        return "prepaid"
    if text in ("postpaid", "arrears", "后付", "期末"):
        return "postpaid"
    return text or "postpaid"


def _normalize_asset_type(value: Any) -> str:
    text = str(value or "").strip().lower()
    if any(keyword in text for keyword in ("店", "铺", "物业", "房", "real")):
        return "real_estate"
    if any(keyword in text for keyword in ("车", "vehicle")):
        return "vehicle"
    if any(keyword in text for keyword in ("电脑", "it", "服务器", "设备")):
        return "it_equipment"
    if any(keyword in text for keyword in ("机器", "machinery")):
        return "machinery"
    return "other" if text else "real_estate"


def _parse_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = (
        str(value)
        .strip()
        .replace(",", "")
        .replace("%", "")
        .replace("¥", "")
        .replace("￥", "")
        .replace("元", "")
    )
    if not text or any(keyword in text for keyword in ("缺失", "待确认", "unknown", "null")):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _coerce_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in ("true", "yes", "y", "1", "是", "有"):
        return True
    if text in ("false", "no", "n", "0", "否", "无"):
        return False
    if any(keyword in text for keyword in ("不行使", "未行使", "不会行使", "不合理确定")):
        return False
    if any(keyword in text for keyword in ("合理确定", "行使")):
        return True
    return default


def _parse_bool_option(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if not text or text in ("无", "否", "false", "no", "n"):
        return False
    if any(keyword in text for keyword in ("不行使", "未行使", "不会行使", "不合理确定")):
        return False
    return any(keyword in text for keyword in ("有", "合理确定", "行使", "true", "yes", "y"))


def _header_index(headers: list[Any]) -> dict[str, int]:
    aliases = {
        "contract_number": ["合同编号", "contract_number", "合同号"],
        "contract_name": ["合同名称", "contract_name", "合同名"],
        "lessee": ["承租方", "法人主体", "legal_entity", "lessee"],
        "lessor": ["出租方", "lessor"],
        "store_name": ["门店/资产名称", "门店名称", "资产名称", "store_name"],
        "store_address": ["门店/资产地址", "门店地址", "资产地址", "store_address"],
        "asset_type": ["资产类型", "资产类别", "asset_type", "asset_category"],
        "currency": ["币种", "currency"],
        "signing_date": ["签约日期", "signing_date"],
        "commencement_date": ["起租日(commencement)", "起租日", "commencement_date", "租赁起始日"],
        "lease_start_date": ["租赁开始日", "lease_start_date"],
        "lease_end_date": ["租赁结束日", "租期结束日", "lease_end_date"],
        "renewal_option": ["续租选择权", "renewal_option"],
        "termination_option": ["终止选择权判断", "终止选择权", "termination_option"],
        "fixed_rent_amount": ["月租金", "固定租金", "fixed_rent_amount"],
        "payment_timing": ["付款时点", "payment_timing"],
        "discount_rate": ["折现率", "discount_rate"],
        "discount_rate_type": ["折现率类型", "discount_rate_type"],
        "lease_scope": ["范围判定(lease_scope)", "lease_scope", "范围判定"],
        "status": ["合同状态", "status"],
    }
    normalized = {str(h or "").strip().lower(): idx for idx, h in enumerate(headers)}
    result: dict[str, int] = {}
    for field, names in aliases.items():
        for name in names:
            key = name.strip().lower()
            if key in normalized:
                result[field] = normalized[key]
                break
    return result


def _get_cell(row: tuple[Any, ...], indexes: dict[str, int], field: str) -> Any:
    idx = indexes.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _as_list(value: Any) -> list:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def _is_excel_content_type(content_type: str) -> bool:
    return content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    )


def _format_excel_workbook_for_llm(file_data: bytes, max_rows: int = 200, max_cols: int = 60) -> str:
    """Unfold workbook cells with coordinates so the LLM can infer non-standard layouts."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
    try:
        text_parts = [
            "Excel workbook cell dump for AI semantic parsing.",
            "The cell coordinates are source evidence. Infer headers and fields from nearby cells, sheet names, and row context.",
        ]
        for ws in wb.worksheets:
            text_parts.append(f"\n## Sheet: {ws.title}")
            max_sheet_row = min(ws.max_row or 0, max_rows)
            max_sheet_col = min(ws.max_column or 0, max_cols)
            for row in ws.iter_rows(
                min_row=1,
                max_row=max_sheet_row,
                min_col=1,
                max_col=max_sheet_col,
            ):
                cells = []
                for cell in row:
                    value = cell.value
                    if value is None or str(value).strip() == "":
                        continue
                    cells.append(f"{cell.coordinate}={_format_excel_date(value)}")
                if cells:
                    text_parts.append(f"Row {row[0].row}: " + " | ".join(cells))
        return "\n".join(text_parts)
    finally:
        wb.close()


def _apply_excel_evidence_safety_checks(contract: dict, file_content: str) -> dict:
    """Guard option booleans against common negated Chinese wording in the source row."""
    contract_number = str(contract.get("contract_number") or "").strip()
    if not contract_number or not file_content:
        return contract

    evidence_line = ""
    for line in file_content.splitlines():
        if contract_number in line:
            evidence_line = line
            break
    if not evidence_line:
        return contract

    if "终止" in evidence_line and any(keyword in evidence_line for keyword in ("不行使", "未行使", "不会行使", "不合理确定")):
        contract["termination_option"] = False
    if "续租" in evidence_line and any(keyword in evidence_line for keyword in ("不续租", "不行使续租", "未行使续租", "不会续租", "不合理确定续租")):
        contract["renewal_option"] = False
    return contract


def _parse_contract_batch_from_excel(file_data: bytes, request: ContractBatchDraftRequest) -> Optional[dict]:
    if not HAS_OPENPYXL:
        return None

    wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
    try:
        validated_contracts: list[ContractBatchItem] = []
        all_warnings: list[str] = []
        all_missing: list[str] = []

        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_row_index = None
            indexes: dict[str, int] = {}
            for idx, row in enumerate(rows[:20]):
                candidate = _header_index(list(row))
                if "contract_number" in candidate and "lessor" in candidate:
                    header_row_index = idx
                    indexes = candidate
                    break

            if header_row_index is None:
                continue

            for row_number, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
                contract_number = str(_get_cell(row, indexes, "contract_number") or "").strip()
                if not contract_number or contract_number.lower() in ("none", "null"):
                    continue

                lease_scope = str(_get_cell(row, indexes, "lease_scope") or "in_scope").strip() or "in_scope"
                contract = {
                    "contract_number": contract_number,
                    "contract_name": str(_get_cell(row, indexes, "contract_name") or contract_number).strip(),
                    "lessee": str(_get_cell(row, indexes, "lessee") or "").strip(),
                    "lessor": str(_get_cell(row, indexes, "lessor") or "").strip(),
                    "store_name": str(_get_cell(row, indexes, "store_name") or "").strip(),
                    "store_address": str(_get_cell(row, indexes, "store_address") or "").strip(),
                    "commencement_date": _format_excel_date(_get_cell(row, indexes, "commencement_date")),
                    "lease_start_date": _format_excel_date(_get_cell(row, indexes, "lease_start_date") or _get_cell(row, indexes, "commencement_date")),
                    "lease_end_date": _format_excel_date(_get_cell(row, indexes, "lease_end_date")),
                    "currency": str(_get_cell(row, indexes, "currency") or "").strip(),
                    "asset_type": _normalize_asset_type(_get_cell(row, indexes, "asset_type")),
                    "fixed_rent_amount": _parse_float(_get_cell(row, indexes, "fixed_rent_amount")),
                    "payment_frequency": "monthly",
                    "payment_timing": _normalize_payment_timing(_get_cell(row, indexes, "payment_timing")),
                    "renewal_option": _parse_bool_option(_get_cell(row, indexes, "renewal_option")),
                    "termination_option": _parse_bool_option(_get_cell(row, indexes, "termination_option")),
                    "cam_amount": 0.0,
                    "service_fee": 0.0,
                    "discount_rate_type": str(_get_cell(row, indexes, "discount_rate_type") or "").strip(),
                    "discount_rate": _parse_float(_get_cell(row, indexes, "discount_rate")),
                    "is_lease": lease_scope != "not_a_lease",
                    "lease_scope": lease_scope,
                    "suggested_scope": lease_scope,
                    "exemption_reason": "",
                    "scope_source": "ledger",
                    "scope_confidence": 0.9 if lease_scope else 0.0,
                    "confidence": 0.9,
                    "warnings": [],
                }

                if not contract["lessee"]:
                    contract["warnings"].append(f"第 {row_number} 行缺少承租方/法人主体")
                if not contract["lessor"]:
                    contract["warnings"].append(f"第 {row_number} 行缺少出租方")

                dr_missing, dr_warnings = _check_discount_rate_missing(contract)
                currency_missing, currency_warnings = _check_currency_missing(contract)
                missing_fields, field_warnings = _check_critical_fields(contract)
                _, scope_warnings = _normalize_lease_scope(contract)

                row_warnings = contract["warnings"] + dr_warnings + currency_warnings + field_warnings + scope_warnings
                all_warnings.extend(row_warnings)
                all_missing.extend(missing_fields)
                if dr_missing:
                    all_missing.append("discount_rate")
                if currency_missing:
                    all_missing.append("currency")

                confidence = 0.9
                if missing_fields or dr_missing or currency_missing:
                    confidence = 0.7
                if float(contract.get("scope_confidence", 0) or 0) < 0.8:
                    confidence = min(confidence, 0.7)

                validated_contracts.append(ContractBatchItem(
                    **{
                        **contract,
                        "confidence": confidence,
                        "missing_fields": missing_fields + (["discount_rate"] if dr_missing else []) + (["currency"] if currency_missing else []),
                        "warnings": row_warnings,
                    }
                ))

        if not validated_contracts:
            return None

        low_confidence_count = sum(1 for c in validated_contracts if (c.confidence or 1.0) < 0.8)
        overall_confidence = sum((c.confidence or 0.8) for c in validated_contracts) / len(validated_contracts)
        unique_missing = sorted(set(all_missing))
        return {
            "task_id": "task_batch_" + request.file_id,
            "file_id": request.file_id,
            "mode": "assist",
            "draft_type": "contract_batch_draft",
            "status": "draft_generated",
            "contracts": validated_contracts,
            "total_count": len(validated_contracts),
            "confidence_scores": {
                "overall": overall_confidence,
                "average_item": overall_confidence,
            },
            "missing_fields": unique_missing,
        "warnings": all_warnings + ["LLM 主解析未能稳定提取合同时，系统已启用 Excel 表格读取兜底；合同草稿必须人工逐条确认后入库"],
            "requires_human_confirmation": overall_confidence < 0.8 or low_confidence_count > 0 or len(unique_missing) > 0,
        }
    finally:
        wb.close()


def _model_to_dict(item: Any) -> dict:
    if hasattr(item, "model_dump"):
        return item.model_dump()
    if hasattr(item, "dict"):
        return item.dict()
    return dict(item)


async def _enrich_excel_contract_batch_with_llm(result: dict) -> dict:
    """Use deterministic Excel extraction as tool output, then ask LLM to review/enrich semantics.

    The spreadsheet reader is responsible for reliable row extraction. The LLM is responsible
    for semantic judgment: IFRS 16 scope, option interpretation, missing-field risk, and
    human-review warnings. If LLM is unavailable, the tool result remains usable.
    """
    contracts = [_model_to_dict(c) for c in result.get("contracts", [])]
    if not contracts:
        return result

    compact_contracts = [
        {
            "contract_number": c.get("contract_number"),
            "contract_name": c.get("contract_name"),
            "lessee": c.get("lessee"),
            "lessor": c.get("lessor"),
            "store_name": c.get("store_name"),
            "asset_type": c.get("asset_type"),
            "commencement_date": c.get("commencement_date"),
            "lease_start_date": c.get("lease_start_date"),
            "lease_end_date": c.get("lease_end_date"),
            "currency": c.get("currency"),
            "fixed_rent_amount": c.get("fixed_rent_amount"),
            "payment_timing": c.get("payment_timing"),
            "renewal_option": c.get("renewal_option"),
            "termination_option": c.get("termination_option"),
            "discount_rate_type": c.get("discount_rate_type"),
            "discount_rate": c.get("discount_rate"),
            "lease_scope": c.get("lease_scope"),
            "missing_fields": c.get("missing_fields"),
            "warnings": c.get("warnings"),
        }
        for c in contracts
    ]

    prompt = f"""
你是租赁管理平台的 AI 录入复核 Agent。系统工具已经从 Excel 台账稳定读取出以下合同草稿。

请不要重新抽取行，也不要删除合同。请基于这些工具结果做语义复核和 IFRS 16 范围判断增强：
1. 保留每个 contract_number。
2. 复核 renewal_option / termination_option 是否被中文表述误判。
3. 复核 asset_type 和 suggested_scope。
4. 如果短期、低价值或非租赁，填写 exemption_reason。
5. 标记缺失字段和人工复核 warning，尤其是折现率缺失、币种缺失、范围判断置信度不足。
6. 返回 JSON，不要 Markdown。

输入合同草稿 JSON:
{json.dumps(compact_contracts, ensure_ascii=False)}

输出格式:
{{
  "contracts": [
    {{
      "contract_number": "LEASE-...",
      "renewal_option": true,
      "termination_option": false,
      "asset_type": "real_estate|vehicle|it_equipment|machinery|other",
      "suggested_scope": "in_scope|short_term_exempt|low_value_exempt|not_a_lease",
      "exemption_reason": "",
      "scope_confidence": 0.0,
      "confidence": 0.0,
      "missing_fields": [],
      "warnings": []
    }}
  ],
  "overall_confidence": 0.0,
  "warnings": []
}}
"""

    try:
        response = await llm_client.chat_completion(
            messages=[
                {
                    "role": "system",
                    "content": "你是租赁管理平台的 AI 录入复核 Agent。工具负责读取表格，你负责语义判断、风险提示和人工复核建议。",
                },
                {"role": "user", "content": prompt},
            ],
            temperature=0.1,
            max_tokens=5000,
            response_format={"type": "json_object"},
        )
        content = response["choices"][0]["message"]["content"]
        parsed = json.loads(content)
    except Exception as exc:
        result["warnings"] = result.get("warnings", []) + [
            f"Excel 台账已由工具稳定解析；LLM 语义复核暂不可用，已保留工具解析结果并要求人工确认。原因: {str(exc)}"
        ]
        result["requires_human_confirmation"] = True
        return result

    enrich_by_number = {
        str(c.get("contract_number", "")).strip(): c
        for c in parsed.get("contracts", [])
        if c.get("contract_number")
    }

    enriched_contracts: list[ContractBatchItem] = []
    for base in contracts:
        contract_number = str(base.get("contract_number", "")).strip()
        enrichment = enrich_by_number.get(contract_number, {})
        merged = {**base}
        for key in (
            "renewal_option",
            "termination_option",
            "asset_type",
            "suggested_scope",
            "exemption_reason",
            "scope_confidence",
            "confidence",
        ):
            if key in enrichment and enrichment[key] not in (None, ""):
                merged[key] = enrichment[key]

        if enrichment.get("suggested_scope"):
            merged["lease_scope"] = enrichment["suggested_scope"]
        merged["scope_source"] = "ai_suggested"
        merged["missing_fields"] = sorted(set((base.get("missing_fields") or []) + (enrichment.get("missing_fields") or [])))
        merged["warnings"] = (base.get("warnings") or []) + (enrichment.get("warnings") or [])
        _, scope_warnings = _normalize_lease_scope(merged)
        merged["warnings"].extend(scope_warnings)
        enriched_contracts.append(ContractBatchItem(**merged))

    overall_confidence = float(parsed.get("overall_confidence") or result.get("confidence_scores", {}).get("overall", 0.8))
    low_confidence_count = sum(1 for c in enriched_contracts if (c.confidence or 1.0) < 0.8)
    missing_fields = sorted({field for c in enriched_contracts for field in (c.missing_fields or [])})

    result["contracts"] = enriched_contracts
    result["total_count"] = len(enriched_contracts)
    result["confidence_scores"] = {
        "overall": overall_confidence,
        "average_item": sum((c.confidence or 0.8) for c in enriched_contracts) / max(len(enriched_contracts), 1),
    }
    result["missing_fields"] = missing_fields
    result["warnings"] = (
        result.get("warnings", [])
        + parsed.get("warnings", [])
        + ["Excel 台账由工具读取，LLM 已完成语义复核；AI Assist Mode: 合同台账草稿需人工逐条确认后入库"]
    )
    result["requires_human_confirmation"] = (
        overall_confidence < 0.8
        or low_confidence_count > 0
        or len(missing_fields) > 0
        or any((c.scope_confidence or 0) < 0.8 for c in enriched_contracts)
    )
    return result


@router.post("/parse/contract-batch", response_model=ContractBatchDraftResponse)
async def parse_contract_batch(request: ContractBatchDraftRequest):
    """
    解析合同台账文件，批量提取多份合同草稿

    适用于：用户上传包含多份合同的 Excel/PDF 台账，AI 提取每份合同的关键字段。

    默认 Assist Mode：
    - AI 只生成草稿，不直接写入正式台账
    - 所有结果需 Editor 逐条确认
    - 关键字段缺失时必须标记人工确认
    """

    if request.mode != "assist":
        raise HTTPException(
            status_code=400,
            detail="当前仅支持 Assist Mode。Auto-Post Mode 需另行配置"
        )

    # Extract text: use provided file_content or download from MinIO + PaddleOCR.
    # For Excel, the workbook reader only unfolds sheet/cell content; LLM remains the
    # primary parser for non-standard layouts.
    file_content = request.file_content
    file_data: Optional[bytes] = None
    is_excel = _is_excel_content_type(request.content_type)
    if not file_content:
        try:
            file_data = download_from_minio("lease-uploads", request.object_name)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"文件下载失败: {str(e)}")

        if is_excel:
            try:
                file_content = _format_excel_workbook_for_llm(file_data)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Excel 文本展开失败: {str(e)}")
        else:
            try:
                file_content = await extract_text(file_data, request.content_type)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"文本提取失败: {str(e)}")

        if len(file_content) > 30000:
            file_content = file_content[:30000] + "\n... (truncated)"

    prompt = f"""
    你是一位专业的 IFRS 16 租赁合同台账解析专家。请从以下合同台账内容中提取每一份合同的字段。

    【重要规则 — 必须遵守】
    1. 台账中可能包含多份合同，请逐条提取
    2. 如果合同中未明确提到折现率/利率，不要猜测，标记为缺失
    3. 如果合同中未明确提到货币（币种），不要猜测，标记为缺失或 unknown
    4. 区分先付租金(prepaid)和后付租金(postpaid)
    5. 区分固定租金和变量租金
    6. 识别租赁成分和非租赁成分(CAM、服务费)
    7. 承租方(lessee)是租赁合同的乙方，即使用物业并支付租金的一方
    8. 出租方(lessor)是租赁合同的甲方，即提供物业并收取租金的一方
    9. 必须做 IFRS 16 范围初判：是否存在已识别资产、承租方是否控制使用、租期是否 ≤12 个月、是否低价值资产。AI 只能建议，不能直接入正式账。
    10. 如果内容来自 Excel，台账可能是非标准排版、多 sheet、多行标题、合并单元格展开后的文本；请按语义理解 sheet 名、标题行、相邻单元格和字段含义，不要依赖固定列名或固定顺序。
    11. 如果出现"法人主体"、"租赁主体"、"承租公司"等字段，通常可作为 lessee/承租方；但仍需结合上下文判断。
    12. 续租/终止选择权必须按否定语义优先判断："不行使"、"未行使"、"不会行使"、"不合理确定"、"无" 均为 false。不要因为文本出现"终止选择权"或"续租选择权"几个字就返回 true。

    合同台账内容:
    {file_content}

    请以 JSON 格式输出，包含以下顶层字段:
    - contracts: 合同列表，每个元素包含:
      - contract_number: 合同编号
      - contract_name: 合同名称
      - lessee: 承租方名称
      - lessor: 出租方名称
      - store_name: 门店/物业名称（如有）
      - store_address: 门店/物业地址（如有）
      - commencement_date: 租赁起始日 (YYYY-MM-DD)
      - lease_start_date: 租赁开始日 (YYYY-MM-DD)
      - lease_end_date: 租期结束日 (YYYY-MM-DD)
      - currency: 币种 (CNY/USD/EUR)。如果未明确提到，返回 null 或空字符串，不要猜测
      - asset_type: 标的资产类型 "real_estate" | "vehicle" | "it_equipment" | "machinery" | "other"
      - fixed_rent_amount: 固定租金金额（仅数字）
      - payment_frequency: 付款频率 (monthly/quarterly/yearly)
      - payment_timing: 付款时点 (prepaid/postpaid)
      - renewal_option: 是否有续租选择权 (true/false)
      - termination_option: 是否有终止选择权 (true/false)
      - cam_amount: 物业管理费 (如有，仅数字)
      - service_fee: 服务费 (如有，仅数字)
      - discount_rate_type: 折现率类型 (如合同中提及)
      - discount_rate: 折现率数值 (如合同中提及)
      - is_lease: 是否构成 IFRS 16 租赁 (true/false)
      - suggested_scope: "in_scope" | "short_term_exempt" | "low_value_exempt" | "not_a_lease"
      - exemption_reason: 范围判定依据，如"租期 10 个月且无续租意图"、"未识别特定资产"
      - scope_confidence: 范围判定置信度 (0-1)
      - confidence: 该份合同识别的置信度 (0-1)
      - missing_fields: 该份合同缺失的字段列表
      - warnings: 该份合同的警告列表
    - total_count: 识别到的合同总数
    - overall_confidence: 总体置信度 (0-1)
    - missing_fields: 全局缺失字段汇总
    - warnings: 全局警告列表
    """

    try:
        response = await llm_client.chat_completion(
            messages=[
                {"role": "system", "content": "你是一位专业的 IFRS 16 租赁合同台账解析专家。请准确提取每份合同字段。如果字段未在合同中出现，不要猜测，标记为缺失。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=8000,
            response_format={"type": "json_object"}
        )

        content = response["choices"][0]["message"]["content"]

        try:
            parsed = json.loads(content)
        except:
            parsed = {
                "contracts": [],
                "total_count": 0,
                "overall_confidence": 0.0,
                "missing_fields": [],
                "warnings": ["解析响应格式异常，请人工检查"]
            }

        contracts_raw = parsed.get("contracts", [])
        if is_excel and len(contracts_raw) == 0 and file_data is not None:
            excel_result = _parse_contract_batch_from_excel(file_data, request)
            if excel_result is not None:
                excel_result["warnings"] = [
                    "LLM 主解析未能从该 Excel 台账稳定提取合同，已启用表格读取兜底；这不是正式入库结果，必须人工逐条确认。"
                ] + excel_result.get("warnings", [])
                excel_result["requires_human_confirmation"] = True
                return excel_result

        validated_contracts = []
        all_warnings = _as_list(parsed.get("warnings", []))
        all_missing = _as_list(parsed.get("missing_fields", []))

        for i, c in enumerate(contracts_raw):
            if is_excel:
                c = _apply_excel_evidence_safety_checks(c, file_content)

            # Validate required fields
            if not c.get("contract_number") or not c.get("lessee") or not c.get("lessor"):
                all_warnings.append(f"第 {i+1} 份合同缺少必要字段 (contract_number/lessee/lessor)，已跳过")
                continue

            # Check discount rate missing
            dr_missing, dr_warnings = _check_discount_rate_missing(c)
            all_warnings.extend(dr_warnings)

            # Check currency missing
            currency_missing, currency_warnings = _check_currency_missing(c)
            all_warnings.extend(currency_warnings)

            # Check critical fields
            missing_fields, field_warnings = _check_critical_fields(c)
            all_warnings.extend(field_warnings)

            # IFRS 16 scope gate suggestion
            _, scope_warnings = _normalize_lease_scope(c)
            all_warnings.extend(scope_warnings)

            contract_confidence = _parse_float(c.get("confidence")) if c.get("confidence") not in (None, "") else 0.8
            if missing_fields:
                contract_confidence = min(contract_confidence, 0.7)
            if _parse_float(c.get("scope_confidence")) < 0.8:
                contract_confidence = min(contract_confidence, 0.7)

            validated_contracts.append(ContractBatchItem(
                contract_number=c.get("contract_number", ""),
                contract_name=c.get("contract_name", ""),
                lessee=c.get("lessee", ""),
                lessor=c.get("lessor", ""),
                store_name=c.get("store_name", ""),
                store_address=c.get("store_address", ""),
                commencement_date=c.get("commencement_date", ""),
                lease_start_date=c.get("lease_start_date", ""),
                lease_end_date=c.get("lease_end_date", ""),
                currency=c.get("currency", ""),
                asset_type=_normalize_asset_type(c.get("asset_type", "real_estate")),
                fixed_rent_amount=_parse_float(c.get("fixed_rent_amount")),
                payment_frequency=c.get("payment_frequency", "monthly"),
                payment_timing=_normalize_payment_timing(c.get("payment_timing", "postpaid")),
                renewal_option=_coerce_bool(c.get("renewal_option"), False),
                termination_option=_coerce_bool(c.get("termination_option"), False),
                cam_amount=_parse_float(c.get("cam_amount")),
                service_fee=_parse_float(c.get("service_fee")),
                discount_rate_type=c.get("discount_rate_type", ""),
                discount_rate=_parse_float(c.get("discount_rate")),
                is_lease=_coerce_bool(c.get("is_lease"), c.get("lease_scope") != "not_a_lease"),
                lease_scope=c.get("lease_scope", "in_scope"),
                suggested_scope=c.get("suggested_scope", c.get("lease_scope", "in_scope")),
                exemption_reason=c.get("exemption_reason", ""),
                scope_source=c.get("scope_source", "ai_suggested"),
                scope_confidence=_parse_float(c.get("scope_confidence")),
                confidence=contract_confidence,
                missing_fields=missing_fields + (["discount_rate"] if dr_missing else []) + (["currency"] if currency_missing else []),
                warnings=dr_warnings + currency_warnings + field_warnings + scope_warnings + _as_list(c.get("warnings", []))
            ))

        overall_confidence = _parse_float(parsed.get("overall_confidence")) if parsed.get("overall_confidence") not in (None, "") else 0.8
        low_confidence_count = sum(1 for c in validated_contracts if (c.confidence or 1.0) < 0.8)

        requires_human = (
            overall_confidence < 0.8 or
            len(validated_contracts) == 0 or
            low_confidence_count > len(validated_contracts) / 2 or
            len(all_missing) > 0
        )

        return {
            "task_id": "task_batch_" + request.file_id,
            "file_id": request.file_id,
            "mode": "assist",
            "draft_type": "contract_batch_draft",
            "status": "draft_generated",
            "contracts": validated_contracts,
            "total_count": len(validated_contracts),
            "confidence_scores": {
                "overall": overall_confidence,
                "average_item": sum((c.confidence or 0.8) for c in validated_contracts) / max(len(validated_contracts), 1),
            },
            "missing_fields": all_missing,
            "warnings": all_warnings + ["AI Assist Mode: 合同台账草稿需人工逐条确认后入库"],
            "requires_human_confirmation": requires_human,
        }
    except Exception as e:
        if is_excel and file_data is not None:
            excel_result = _parse_contract_batch_from_excel(file_data, request)
            if excel_result is not None:
                excel_result["warnings"] = [
                    f"LLM 主解析暂不可用或返回异常，已启用 Excel 表格读取兜底；合同草稿必须人工逐条确认。原因: {str(e)}"
                ] + excel_result.get("warnings", [])
                excel_result["requires_human_confirmation"] = True
                return excel_result
        raise HTTPException(status_code=500, detail=f"批量解析失败: {str(e)}")
