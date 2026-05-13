"""Document text extraction service.

Primary: PaddleOCR-VL-1.5 (AI Studio 异步 API) — 支持 PDF、scan copy、图片
Fallback: PyMuPDF (纯文本 PDF)、openpyxl (Excel)
"""

import logging
import io
from typing import Optional

from app.services.paddleocr import get_paddleocr_client

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


async def extract_text_from_pdf_with_paddleocr(data: bytes, filename: str = "document.pdf") -> Optional[str]:
    """使用 PaddleOCR-VL 解析 PDF（含 scan copy）。

    Args:
        data: PDF 文件二进制内容
        filename: 文件名（用于 PaddleOCR 提交）

    Returns:
        Markdown 文本，或 None 表示 PaddleOCR 不可用
    """
    client = get_paddleocr_client()
    if not client.is_available():
        logger.info("PaddleOCR 未启用，将使用 fallback")
        return None

    try:
        markdown = await client.extract_text_from_file(
            file_data=data,
            filename=filename,
            content_type="application/pdf",
        )

        if markdown and len(markdown.strip()) > 50:
            logger.info(f"PaddleOCR 解析成功，文本长度: {len(markdown)}")
            return markdown
        else:
            logger.warning(f"PaddleOCR 返回文本过短 ({len(markdown or '')} 字符)，尝试 fallback")
            return None

    except Exception as e:
        logger.warning(f"PaddleOCR 解析失败，使用 fallback: {e}")
        return None


async def extract_text_from_pdf_with_paddleocr_url(file_url: str) -> Optional[str]:
    """使用 PaddleOCR-VL 通过 URL 解析 PDF。

    Args:
        file_url: 可公开访问的文件 URL

    Returns:
        Markdown 文本，或 None 表示 PaddleOCR 不可用
    """
    client = get_paddleocr_client()
    if not client.is_available():
        logger.info("PaddleOCR 未启用，将使用 fallback")
        return None

    try:
        markdown = await client.extract_text_from_url(file_url)
        if markdown and len(markdown.strip()) > 50:
            logger.info(f"PaddleOCR URL 解析成功，文本长度: {len(markdown)}")
            return markdown
        else:
            logger.warning(f"PaddleOCR URL 返回文本过短 ({len(markdown or '')} 字符)，尝试 fallback")
            return None

    except Exception as e:
        logger.warning(f"PaddleOCR URL 解析失败，使用 fallback: {e}")
        return None


async def extract_text_from_image_with_paddleocr(data: bytes, filename: str = "image.jpg", content_type: str = "image/jpeg") -> Optional[str]:
    """使用 PaddleOCR 解析图片（scan copy、照片）。

    Args:
        data: 图片文件二进制内容
        filename: 文件名
        content_type: MIME 类型

    Returns:
        Markdown 文本，或 None 表示 PaddleOCR 不可用
    """
    client = get_paddleocr_client()
    if not client.is_available():
        logger.info("PaddleOCR 未启用，无法解析图片")
        return None

    try:
        markdown = await client.extract_text_from_file(
            file_data=data,
            filename=filename,
            content_type=content_type,
        )
        logger.info(f"PaddleOCR 图片解析成功，文本长度: {len(markdown or '')}")
        return markdown

    except Exception as e:
        logger.warning(f"PaddleOCR 图片解析失败: {e}")
        return None


async def extract_text_from_image_with_paddleocr_url(file_url: str) -> Optional[str]:
    """使用 PaddleOCR 通过 URL 解析图片。

    Args:
        file_url: 可公开访问的图片 URL

    Returns:
        Markdown 文本，或 None 表示 PaddleOCR 不可用
    """
    client = get_paddleocr_client()
    if not client.is_available():
        logger.info("PaddleOCR 未启用，无法解析图片")
        return None

    try:
        markdown = await client.extract_text_from_url(file_url)
        logger.info(f"PaddleOCR 图片 URL 解析成功，文本长度: {len(markdown or '')}")
        return markdown

    except Exception as e:
        logger.warning(f"PaddleOCR 图片 URL 解析失败: {e}")
        return None


def extract_text_from_pdf_fallback(data: bytes, max_pages: int = 50) -> str:
    """Fallback: 使用 PyMuPDF 提取纯文本 PDF。"""
    if not HAS_PYMUPDF:
        raise RuntimeError("PyMuPDF not installed")

    doc = fitz.open(stream=data, filetype="pdf")
    text_parts = []

    for page_num in range(min(len(doc), max_pages)):
        page = doc[page_num]
        text = page.get_text()
        if text.strip():
            text_parts.append(f"--- Page {page_num + 1} ---\n{text}")

    doc.close()
    return "\n\n".join(text_parts)


def extract_text_from_excel(data: bytes, max_rows: int = 1000) -> str:
    """Extract text from Excel bytes as markdown table."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    text_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"### Sheet: {sheet_name}")

        rows = []
        for row in ws.iter_rows(values_only=True, max_row=max_rows):
            row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
            rows.append(row_str)

        if rows:
            if len(rows) > 1:
                header = rows[0]
                separator = " | ".join(["---"] * len(header.split(" | ")))
                text_parts.append(header)
                text_parts.append(separator)
                text_parts.extend(rows[1:])
            else:
                text_parts.extend(rows)

        text_parts.append("")

    wb.close()
    return "\n".join(text_parts)


async def extract_text(data: bytes, content_type: str, file_url: Optional[str] = None) -> str:
    """提取文件文本。优先使用 PaddleOCR，失败时 fallback 到本地提取。

    Args:
        data: 文件二进制内容
        content_type: MIME 类型
        file_url: 可选的公开访问 URL（用于 PaddleOCR fileUrl 方式，优先于 base64）

    Returns:
        提取的文本内容
    """
    if content_type == "application/pdf":
        # 优先 PaddleOCR（支持 scan copy + 复杂版面）
        # 优先使用 URL 方式（更稳定），回退到 base64 方式
        if file_url:
            paddleocr_text = await extract_text_from_pdf_with_paddleocr_url(file_url)
            if paddleocr_text:
                return paddleocr_text

        paddleocr_text = await extract_text_from_pdf_with_paddleocr(data)
        if paddleocr_text:
            return paddleocr_text

        # Fallback: PyMuPDF 纯文本提取
        logger.info("使用 PyMuPDF fallback 提取 PDF 文本")
        return extract_text_from_pdf_fallback(data)

    elif content_type in ("image/jpeg", "image/png", "image/tiff", "image/bmp", "image/webp"):
        # 图片只能用 PaddleOCR
        if file_url:
            paddleocr_text = await extract_text_from_image_with_paddleocr_url(file_url)
            if paddleocr_text:
                return paddleocr_text

        paddleocr_text = await extract_text_from_image_with_paddleocr(data)
        if paddleocr_text:
            return paddleocr_text
        raise RuntimeError(f"PaddleOCR 不可用，无法解析图片 ({content_type})。请配置 PADDLEOCR_ACCESS_TOKEN。")

    elif content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        # Excel 用 openpyxl，不走 OCR
        return extract_text_from_excel(data)

    else:
        raise ValueError(f"不支持的文件类型: {content_type}")