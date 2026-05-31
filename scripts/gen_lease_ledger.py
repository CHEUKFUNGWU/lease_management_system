from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

headers = [
    "合同编号", "合同名称", "法人主体", "资产编号", "门店/资产名称", "出租方",
    "资产类别", "币种", "签约日期", "起租日(Commencement)", "租赁结束日",
    "不可撤销期(月)", "续租选择权", "终止选择权判断", "月租金", "付款时点",
    "折现率", "折现率类型", "范围判定(lease_scope)", "合同状态",
]

# 范围判定取值: in_scope / short_term_exempt / low_value_exempt / not_a_lease
rows = [
    ["LEASE-2024-001","南京东路旗舰店租约","零售集团上海公司","STORE-SH-001","南京东路旗舰店","上海置地商业管理有限公司","商业铺位","CNY","2023-12-15","2024-01-01","2026-12-31",36,"5年续租选择权(合理确定行使)","不行使提前终止",50000,"后付",0.0525,"增量借款利率","in_scope","approved"],
    ["LEASE-2024-002","王府井标准店租约","零售集团北京公司","STORE-BJ-002","北京王府井标准店","王府井集团股份有限公司","商业铺位","CNY","2024-01-10","2024-02-01","2029-01-31",60,"无","不行使提前终止",42000,"先付",0.0550,"增量借款利率","in_scope","approved"],
    ["LEASE-2024-003","天河城购物中心店租约","零售集团广州公司","STORE-GZ-003","广州天河城店","天河城集团有限公司","商业铺位","CNY","2024-02-20","2024-03-01","2027-02-28",36,"3年续租选择权(未确定行使)","可能行使提前终止",38000,"后付",0.0500,"增量借款利率","in_scope","approved"],
    ["LEASE-2024-004","铜锣湾旗舰店租约","零售集团香港公司","STORE-HK-004","香港铜锣湾旗舰店","希慎兴业有限公司","商业铺位","HKD","2024-01-05","2024-01-15","2030-01-14",72,"无","不行使提前终止",180000,"先付",0.0480,"增量借款利率","in_scope","pending_approval"],
    ["LEASE-2024-005","春熙路步行街店租约","零售集团成都公司","STORE-CD-005","成都春熙路店","成都商业地产有限公司","商业铺位","CNY","2024-03-01","2024-04-01","2028-03-31",48,"2年续租选择权(合理确定行使)","不行使提前终止",35000,"后付",None,"待人工确认","in_scope","draft"],
    ["LEASE-2024-006","湖滨银泰in77店租约","零售集团杭州公司","STORE-HZ-006","杭州湖滨银泰店","银泰商业集团","商业铺位","CNY","2024-02-15","2024-03-15","2029-03-14",60,"无","不行使提前终止",40000,"后付",0.0535,"增量借款利率","in_scope","approved"],
    ["LEASE-2024-007","光谷步行街店租约","零售集团武汉公司","STORE-WH-007","武汉光谷店","武汉光谷商业管理","商业铺位","CNY","2024-04-01","2024-05-01","2027-04-30",36,"3年续租选择权(未确定行使)","不行使提前终止",28000,"后付",0.0560,"增量借款利率","in_scope","pending_approval"],
    ["LEASE-2024-008","赛格购物中心店租约","零售集团西安公司","STORE-XA-008","西安赛格店","赛格国际购物中心","商业铺位","CNY","2024-03-20","2024-04-15","2028-04-14",48,"无","可能行使提前终止",26000,"后付",0.0570,"增量借款利率","in_scope","approved"],
    ["LEASE-2024-009","华东中央仓库租约","零售集团上海公司","WH-SH-009","上海青浦中央仓库","普洛斯物流地产","仓库","CNY","2024-01-20","2024-02-01","2031-01-31",84,"5年续租选择权(合理确定行使)","不行使提前终止",120000,"后付",0.0510,"增量借款利率","in_scope","approved"],
    ["LEASE-2024-010","华南配送中心租约","零售集团广州公司","WH-GZ-010","东莞华南配送中心","宝湾物流控股","仓库","CNY","2024-02-10","2024-03-01","2030-02-28",72,"无","不行使提前终止",95000,"后付",0.0515,"增量借款利率","in_scope","approved"],
    ["LEASE-2024-011","集团总部办公室租约","零售集团总公司","OFF-SH-011","上海陆家嘴总部办公室","陆家嘴金融贸易区开发","办公室","CNY","2023-11-01","2024-01-01","2028-12-31",60,"5年续租选择权(合理确定行使)","不行使提前终止",220000,"先付",0.0490,"增量借款利率","in_scope","approved"],
    ["LEASE-2024-012","配送车队租赁(10辆)","零售集团上海公司","VEH-012","城配货车车队(10辆)","中车租赁有限公司","车辆","CNY","2024-01-15","2024-02-01","2027-01-31",36,"无","不行使提前终止",45000,"先付",0.0600,"承租人增量借款利率","in_scope","approved"],
    ["LEASE-2024-013","冷链运输车租赁","零售集团广州公司","VEH-013","冷链运输车(3辆)","狮桥融资租赁","车辆","CNY","2024-03-01","2024-03-15","2028-03-14",48,"购买选择权(合理确定行使)","不行使提前终止",28000,"后付",0.0620,"合同内含利率","in_scope","pending_approval"],
    ["LEASE-2024-014","双十一临时快闪店","零售集团上海公司","POP-014","上海静安嘉里快闪店","嘉里建设","商业铺位","CNY","2024-09-20","2024-10-20","2025-01-19",3,"无","无","80000","先付",None,"豁免-无需贴现","short_term_exempt","approved"],
    ["LEASE-2024-015","春节临时周转仓","零售集团北京公司","POP-015","北京大兴临时仓","顺丰物流","仓库","CNY","2023-12-25","2024-01-15","2024-03-14",2,"无","无","30000","后付",None,"豁免-无需贴现","short_term_exempt","approved"],
    ["LEASE-2024-016","办公笔记本租赁(20台)","零售集团总公司","IT-016","联想笔记本电脑(20台)","联想租赁服务","IT设备","CNY","2024-01-10","2024-02-01","2027-01-31",36,"无","不行使提前终止","6000","后付",None,"豁免-低价值资产","low_value_exempt","approved"],
    ["LEASE-2024-017","门店打印复印设备租赁","零售集团上海公司","IT-017","打印机/复印机(15台)","震旦办公设备租赁","IT设备","CNY","2024-02-01","2024-02-15","2027-02-14",36,"无","不行使提前终止","4500","后付",None,"豁免-低价值资产","low_value_exempt","approved"],
    ["LEASE-2024-018","茶水间咖啡机租赁","零售集团总公司","EQ-018","商用咖啡机(8台)","雀巢专业餐饮","设备","CNY","2024-03-01","2024-03-10","2026-03-09",24,"无","不行使提前终止","2800","后付",None,"豁免-低价值资产","low_value_exempt","approved"],
    ["LEASE-2024-019","云服务器IDC托管服务","零售集团总公司","SVC-019","阿里云IDC托管服务","阿里云计算有限公司","IT服务","CNY","2024-01-01","2024-01-01","2026-12-31",36,"无","无","35000","后付",None,"非租赁-无已识别资产","not_a_lease","approved"],
    ["LEASE-2024-020","第三方仓储外包服务","零售集团广州公司","SVC-020","佛山3PL仓储外包","百世供应链","物流服务","CNY","2024-02-01","2024-02-01","2025-01-31",12,"无","无","60000","后付",None,"非租赁-承租人无控制权","not_a_lease","pending_approval"],
]

wb = Workbook()
ws = wb.active
ws.title = "合同台账"

FONT = "Arial"
header_fill = PatternFill("solid", start_color="1F4E78")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
cell_font = Font(name=FONT, size=10)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

scope_fill = {
    "in_scope": PatternFill("solid", start_color="E2EFDA"),
    "short_term_exempt": PatternFill("solid", start_color="FFF2CC"),
    "low_value_exempt": PatternFill("solid", start_color="FCE4D6"),
    "not_a_lease": PatternFill("solid", start_color="F2F2F2"),
}
status_color = {
    "approved": "008000", "pending_approval": "BF8F00",
    "draft": "808080", "rejected": "FF0000",
}

ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

scope_idx = headers.index("范围判定(lease_scope)") + 1
status_idx = headers.index("合同状态") + 1
rate_idx = headers.index("折现率") + 1
rent_idx = headers.index("月租金") + 1

for r in rows:
    ws.append(r)
    rownum = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=rownum, column=c)
        cell.font = cell_font
        cell.border = border
        cell.alignment = Alignment(vertical="center")
    # 月租金千分位
    ws.cell(row=rownum, column=rent_idx).number_format = '#,##0'
    ws.cell(row=rownum, column=rent_idx).alignment = Alignment(horizontal="right", vertical="center")
    # 折现率百分比
    rc = ws.cell(row=rownum, column=rate_idx)
    if rc.value is None:
        rc.value = "缺失-待确认"
        rc.font = Font(name=FONT, size=10, color="FF0000")
        rc.alignment = Alignment(horizontal="center", vertical="center")
    else:
        rc.number_format = '0.00%'
        rc.alignment = Alignment(horizontal="center", vertical="center")
    # 范围判定底色
    sc = ws.cell(row=rownum, column=scope_idx)
    sc.fill = scope_fill.get(sc.value, PatternFill())
    sc.alignment = Alignment(horizontal="center", vertical="center")
    # 状态字体色
    stc = ws.cell(row=rownum, column=status_idx)
    stc.font = Font(name=FONT, size=10, bold=True, color=status_color.get(stc.value, "000000"))
    stc.alignment = Alignment(horizontal="center", vertical="center")

widths = [16,24,18,14,22,26,10,8,12,18,14,12,26,18,12,10,12,18,22,16]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
ws.row_dimensions[1].height = 32

# 图例 sheet
leg = wb.create_sheet("字段说明")
legend = [
    ["字段说明 / 测试数据图例", ""],
    ["", ""],
    ["范围判定(lease_scope)", "含义"],
    ["in_scope", "适用 IFRS 16 资本化，需算 ROU + 租赁负债（绿色底）"],
    ["short_term_exempt", "短期租赁豁免(≤12月)，直线法费用化，不上表（黄色底）"],
    ["low_value_exempt", "低价值资产豁免，直线法费用化，不上表（橙色底）"],
    ["not_a_lease", "非租赁合同(无已识别资产/无控制权)，不入 IFRS 16（灰色底）"],
    ["", ""],
    ["付款时点", "先付=期初支付(影响ROU初始成本) / 后付=期末支付(纳入贴现)"],
    ["折现率", "缺失时标红，触发 human-in-the-loop 人工确认，AI 不得猜测"],
    ["合同状态", "approved(绿) / pending_approval(黄) / draft(灰) / rejected(红)"],
    ["", ""],
    ["说明", "本表为测试数据，共20条，覆盖门店/仓库/办公室/车辆/IT设备/服务等资产类型"],
    ["", "币种含 CNY 与 HKD；含1条折现率缺失场景(LEASE-2024-005)"],
]
for row in legend:
    leg.append(row)
leg["A1"].font = Font(name=FONT, bold=True, size=12)
for r in range(3, 4):
    for c in range(1, 3):
        leg.cell(row=r, column=c).font = Font(name=FONT, bold=True)
leg.column_dimensions["A"].width = 26
leg.column_dimensions["B"].width = 60
for row in leg.iter_rows():
    for cell in row:
        if not (cell.font and cell.font.bold):
            cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

wb.save("合同台账_测试数据.xlsx")
print("saved")
